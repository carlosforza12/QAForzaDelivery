import re
import random
import os
from pathlib import Path
import pyodbc
import openpyxl
from openpyxl.utils import column_index_from_string
import allure
from playwright.sync_api import Page, expect

class ForzaPage:
    def __init__(self, page: Page, db_connection_string: str = ""):
        self.page = page
        self.db_connection_string = db_connection_string #
        
        # Variables de estado
        self.ruta_excel: str = ""
        self.hoja: str = ""
        self.columna: str = ""
        self.url_actual: str = ""
        self.token: str = ""
        self.entorno: str = ""

    def _take_screenshot(self, name: str = "screenshot"):
        try:
            screenshot = self.page.screenshot(type="png")
            allure.attach(screenshot, name=name, attachment_type=allure.attachment_type.PNG)
        except Exception as e:
            print(f"Error taking screenshot: {e}")
    
    @allure.step("Navegar a la URL: '{url}'")
    def go_to_page_web(self, url: str):
        self.url_actual = url
        self.page.goto(url, timeout=120000)
        self._take_screenshot("navigation")

    @allure.step("Validar que el título sea: '{titulo_esperado}'")
    def assert_title(self, titulo_esperado: str):
        assert self.page.title() == titulo_esperado, f"Error: Título esperado '{titulo_esperado}', actual '{self.page.title()}'"
        self._take_screenshot("title_validation")

    @allure.step("Seleccionar país: '{country}'")
    def select_country(self, country: str):
        patron_regex = re.compile(f"^{re.escape(country)}$")
        self.page.locator("div").filter(has_text=patron_regex).get_by_role("emphasis").click(timeout=30000)
        self.page.wait_for_load_state("load")
        self._take_screenshot("country_selection")

    @allure.step("Seleccionar país en App Courier: '{country}'")
    def select_country_courier(self, country: str):
        country_text = self.page.get_by_text(country, exact=True)
        if country_text.count() > 0 and country_text.first.is_visible():
            self._take_screenshot("courier_country_selection")
            return

        try:
            self.page.get_by_role("combobox").select_option(label=country, timeout=10000)
        except Exception as first_error:
            try:
                self.page.locator("select").first.select_option(label=country, timeout=5000)
            except Exception:
                try:
                    self.page.get_by_role("combobox").first.click(timeout=5000)
                    self.page.get_by_text(country, exact=True).click(timeout=5000)
                except Exception:
                    if country_text.count() > 0:
                        return
                    raise first_error
        self._take_screenshot("courier_country_selection")

    # ==============================================================================
    # FLUJOS DE LOGIN
    # ==============================================================================

    @allure.step("Login Portal Web - Correo: '{correo}'")
    def login(self, correo: str, passw: str):
        self.page.get_by_role("textbox", name="Ingresar Usuario").click()
        self.page.get_by_role("textbox", name="Ingresar Usuario").fill(correo)
        
        self.page.get_by_role("textbox", name="Ingresar Contraseña").click()
        self.page.get_by_role("textbox", name="Ingresar Contraseña").fill(passw)
        
        self.page.get_by_role("button", name="Iniciar sesión").click()
        self.page.wait_for_url(f"**/design/dashboard", wait_until="networkidle")
        self._take_screenshot("login_success")

    @allure.step("Login Corporativo - Código: '{codigo}', Usuario: '{usuario}'")
    def login_corp(self, codigo: str, usuario: str, passw: str):
        self.page.get_by_role("button", name="Usuario Corporativo").click()
        
        self.page.get_by_placeholder("Ingresar Código").click()
        self.page.get_by_placeholder("Ingresar Código").fill(codigo)
        
        self.page.get_by_role("textbox", name="Ingresar Usuario").click()
        self.page.get_by_role("textbox", name="Ingresar Usuario").fill(usuario)
        
        self.page.get_by_role("textbox", name="Ingresar Contraseña").click()
        self.page.get_by_role("textbox", name="Ingresar Contraseña").fill(passw)
        
        self.page.get_by_role("button", name="Iniciar sesión").click()
        self.page.wait_for_url(f"**/historico", wait_until="networkidle")
        self._take_screenshot("login_corp_success")

    @allure.step("Login Express Center - Estación: '{estacion}', Correo: '{correo}'")
    def login_exec(self, estacion: str, correo: str, passw: str):
        self.page.get_by_role("button", name="Usuario Express Center").click()
        self.page.get_by_role("button", name="Selecciona tu estación").click()
        
        self.page.get_by_role("searchbox", name="search text").click()
        self.page.get_by_role("searchbox", name="search text").fill(estacion)
        self.page.get_by_text(estacion, exact=True).click()
        
        self.page.get_by_role("textbox", name="Ingresar Usuario").click()
        self.page.get_by_role("textbox", name="Ingresar Usuario").fill(correo)
        
        self.page.get_by_role("textbox", name="Ingresar Contraseña").click()
        self.page.get_by_role("textbox", name="Ingresar Contraseña").fill(passw)
        
        self.page.get_by_role("button", name="Iniciar sesión").click()
        self.page.wait_for_url(f"**/historico", wait_until="networkidle")
        self._take_screenshot("login_exec_success")

    @allure.step("Login App Courier - Teléfono: '{telefono}'")
    def login_courier_app(self, telefono: str, entorno: str = None):
        self.page.wait_for_timeout(1000)
        input_phone = self.page.locator("#inputPhone")
        if input_phone.count() > 0:
            input_phone.fill(telefono)
        else:
            self.page.locator("input:not([type='hidden'])").last.fill(telefono)
        self.page.wait_for_timeout(1000)
        
        self.page.get_by_role("button", name="Autenticar").click()
        self.page.get_by_text("Autenticación").wait_for()
        
        # Use provided entorno or instance entorno
        entorno_a_usar = entorno if entorno is not None else self.entorno
        self.token = self.obtener_token_db(telefono, entorno_a_usar)
        self.ingresar_token(self.token)
        self._take_screenshot("login_courier_success")

    @allure.step("Ingresar Token OTP")
    def ingresar_token(self, token: str):
        if not token or not token.strip():
            raise ValueError("Token inválido o vacío.")
            
        for i, digito in enumerate(token):
            self.page.locator(f"#inputToken{i + 1}").fill(digito)
            
        self.page.get_by_text("Iniciar sesión").click()
        self.page.wait_for_load_state("networkidle")
        self.page.wait_for_timeout(2000)
        self._take_screenshot("token_entered")

    # ==============================================================================
    # ACCIONES DE NAVEGACIÓN INTERNA
    # ==============================================================================

    @allure.step("Seleccionar opción tipo de servicio: '{tipo_servicio}'")
    def opcion_tipo_servicio(self, tipo_servicio: str):
        self.page.get_by_title(tipo_servicio).click()
        self._take_screenshot("service_type_selected")

    @allure.step("Seleccionar opción Crear Guías")
    def opcion_crear_guias(self):
        self.page.get_by_role("link", name="chevron forward Crear Guías").click()
        self.page.wait_for_timeout(1000)
        self._take_screenshot("create_guides_section")

    # ==============================================================================
    # FLUJO PRINCIPAL: CREACIÓN DE GUÍAS
    # ==============================================================================

    @allure.step("Creación de {cantidad} guías con parámetros específicos")
    def creacion_guias(self, cantidad: int, direccion): # 'direccion' es el dataclass de test_forza_steps.py
        modal_ya_fue_cerrado = False
        is_collet = str(direccion.collet).lower() == 'true'
        is_nueva_origen = str(direccion.direccion_nueva_origen).lower() == 'true'
        is_nueva_destino = str(direccion.direccion_nueva_destino).lower() == 'true'
        is_std = direccion.tipo_guia == "Servicio Estándar"
        
        for i in range(cantidad):
            print(f"Registrando solicitud #{i + 1}")
            monto_aleatorio = 0
            self.page.wait_for_timeout(1000)
            self.page.get_by_title(direccion.tipo_guia).click()

            if is_nueva_origen:
                # Aquí puedes poner la URL mapeada desde tu configuración
                self.page.wait_for_url("**/opciones-guias", wait_until="networkidle", timeout=30000)
                self.page.get_by_text("Crear dirección de origen").click()
                self.crear_nueva_direccion_origen()

            self.page.locator("#step5 #buttonOpenSelect").click()

            if is_nueva_destino:
                self.page.locator("//strong[contains(text(), 'Crear nueva dirección')]").click()
                self.page.wait_for_timeout(1000)
                self.crear_nueva_direccion_destino()
            else:
                self.page.get_by_text(direccion.nombre_direccion).click()
                self.page.wait_for_timeout(1000)

            if is_collet:
                self.page.get_by_label("Cobro contra entrega").check()
            else:
                self.page.get_by_label("Cobro a mi cuenta").check()

            self.page.locator("#step8 input[type='text']").fill("QA002")
            self.page.wait_for_timeout(1000)

            if is_std:
                self.page.get_by_text("Agregar paquete").click()
            else:
                monto_aleatorio = random.randint(5000, 20000)
                self.page.locator("#step19 input[type='text']").fill(str(monto_aleatorio))
                self.page.locator("#step20 #buttonOpenSelect").click()
                # ¡AQUÍ ESTÁ LA MAGIA! Reemplazamos el string quemado por direccion.tarjeta
                self.page.get_by_text(direccion.tarjeta).click()
                self.page.get_by_text("Agregar paquete").click()
                self.page.wait_for_timeout(1000)

            if not modal_ya_fue_cerrado:
                self.page.locator("#packageInformationModal button").click()
                modal_ya_fue_cerrado = True

            self.page.wait_for_timeout(1000)
            self.page.get_by_title("Finalizar y pagar").click()
            self.page.wait_for_timeout(1000)
            self.page.get_by_role("button", name="Si").click()
            self._take_screenshot(f"guide_{i+1}_confirm")

            if monto_aleatorio > 10000:
                self.page.wait_for_timeout(1000)
                self.page.get_by_role("button", name="Si").click()

            if is_collet:
                self.page.get_by_text("Finalizar", exact=True).click()
                self.page.locator("div").filter(has_text=re.compile(r"^¿Cómo realizarás tu recolección\?$")).locator("button").click()
                self.page.get_by_title("Mis Envíos").click()
                self.page.wait_for_timeout(1000)
                
                numero_guia = self.page.locator("(//table//tr[1]//strong[contains(text(), 'FD')])[1]").inner_text()
                print(f"Número de guía: {numero_guia}")
                
                if is_std:
                    self.guardar_guia_en_archivo("guias_Collet_STD", "STD", numero_guia)
                else:
                    self.guardar_guia_en_archivo("guias_Collet_COD", "COD", numero_guia)
                self._take_screenshot(f"guide_{i+1}_created")
            else:
                self.page.wait_for_timeout(1000)
                self.page.get_by_label("CF").check()
                self.page.get_by_label("Tarjeta").check()
                self.page.get_by_text("Pagar Carrito").click()
                self.page.wait_for_timeout(1000)
                self.page.wait_for_load_state("load")

                frame = self.page.frame_locator("#RequestFAC")
                texto_frame = frame.get_by_role("paragraph").inner_text()

                match = re.search(r"FD\d+", texto_frame)
                if match:
                    guia = match.group(0)
                    if is_std:
                        self.guardar_guia_en_archivo("guias_STD_Prepagada", "STD", guia)
                    else:
                        self.guardar_guia_en_archivo("guias_Collet_COD_Prepagada", "COD", guia)
                    print(f"Guía encontrada: {guia}")
                else:
                    print("No se encontró una guía.")
                self._take_screenshot(f"guide_{i+1}_prepaid")

                # TRADUCCIÓN DEL PATRÓN "Task" DE C# A PYTHON (Eventos asíncronos en API Sync)
                with self.page.expect_popup() as popup_info, self.page.expect_download() as download_info:
                    self.page.locator("#confirmPaymentModal button").click()
                
                popup = popup_info.value
                download = download_info.value

                self.page.locator("div").filter(has_text=re.compile(r"^¿Cómo realizarás tu recolección\?$")).locator("button").click()

    # ==============================================================================
    # DIRECCIONES Y AYUDANTES DE FORMULARIOS
    # ==============================================================================

    @allure.step("Llenar formulario de nueva dirección destino")
    def crear_nueva_direccion_destino(self):
        nombre = self.generar_nombre_qa()
        telefono = self.generar_telefono()

        self.page.get_by_placeholder("Ingresar Nombre").nth(1).click()
        self.page.get_by_placeholder("Ingresar Nombre").nth(1).fill(nombre)

        self.page.get_by_placeholder("Lugar en la ciudad").nth(1).click()
        self.page.locator("//li[normalize-space(.)='Gimnasio']").nth(1).click()

        self.page.get_by_placeholder("Seleccione Poblado").nth(1).click()
        self.page.locator("//li[normalize-space(.)='Cahabon, Santa Maria Cahabon, Alta Verapaz']").nth(1).click()

        self.page.get_by_placeholder("Ingresar Teléfono").nth(1).click()
        self.page.get_by_placeholder("Ingresar Teléfono").nth(1).fill(telefono)

        self.page.get_by_placeholder("Ingrese la dirección exacta").nth(1).click()
        self.page.get_by_placeholder("Ingrese la dirección exacta").nth(1).fill("Cahabón, Guatemala")

        self.page.locator("//input[@placeholder='Ingrese el nombre de la persona que recibirá el envío' and @type='text']").click()
        self.page.locator("//input[@placeholder='Ingrese el nombre de la persona que recibirá el envío' and @type='text']").fill("QA Demo")

        # Forzar validaciones (Presionar espacio)
        self.page.get_by_placeholder("Ingresar Nombre").nth(1).press(" ")
        self.page.get_by_placeholder("Ingresar Teléfono").nth(1).press(" ")
        self.page.locator("//input[@placeholder='Ingrese el nombre de la persona que recibirá el envío' and @type='text']").press(" ")
        self.page.get_by_placeholder("Ingrese la dirección exacta").nth(1).press(" ")

        self.page.locator("//button[normalize-space(.)='Guardar']").nth(3).click()
        self._take_screenshot("new_destination_address")

    @allure.step("Llenar formulario de nueva dirección origen")
    def crear_nueva_direccion_origen(self):
        nombre = self.generar_nombre_qa()
        telefono = self.generar_telefono()

        self.page.get_by_placeholder("Ingresar Nombre").nth(0).click()
        self.page.get_by_placeholder("Ingresar Nombre").nth(0).fill(nombre)

        self.page.get_by_placeholder("Lugar en la ciudad").nth(0).click()
        self.page.locator("//li[normalize-space(.)='Gimnasio']").nth(0).click()

        self.page.get_by_placeholder("Seleccione Poblado").nth(0).click()
        self.page.locator("//li[normalize-space(.)='Cahabon, Santa Maria Cahabon, Alta Verapaz']").nth(0).click()

        self.page.get_by_placeholder("Ingresar Teléfono").nth(0).click()
        self.page.get_by_placeholder("Ingresar Teléfono").nth(0).fill(telefono)

        self.page.get_by_placeholder("Ingrese la dirección exacta").nth(0).click()
        self.page.get_by_placeholder("Ingrese la dirección exacta").nth(0).fill("Cahabón, Guatemala")

        self.page.get_by_placeholder("Ingrese ubicación ej. zona 10").click()
        self.page.get_by_placeholder("Ingrese ubicación ej. zona 10").fill("Ciudad de guatemala")

        input_ub = self.page.get_by_placeholder("Ingrese ubicación ej. zona 10")
        input_ub.press("Backspace")

        self.page.get_by_text("Ciudad de Guatemala", exact=True).click()

        self.page.get_by_placeholder("Ingresar Nombre").nth(0).press(" ")
        self.page.get_by_placeholder("Ingresar Teléfono").nth(0).press(" ")
        self.page.locator("//input[@placeholder='Ingrese el nombre de la persona que recibirá el envío' and @type='text']").press(" ")
        self.page.get_by_placeholder("Ingrese la dirección exacta").nth(0).press(" ")

        self.page.locator(".form-group > .btn").first.click()
        self._take_screenshot("new_origin_address")

    # ==============================================================================
    # PROCESAMIENTO DE EXCEL (Equivalente a ClosedXML)
    # ==============================================================================

    def ruta_excel_obtener(self, ruta: str):
        self.ruta_excel = ruta

    def hoja_excel(self, hoja: str):
        self.hoja = hoja

    def columna_excel(self, col: str):
        self.columna = col

    @allure.step("Ingresar lote de recolección desde Excel (Inicio: {inicial}, Final: {final}, Piezas: {cantidad_piezas})")
    def ingresar_a_recoleccion(self, inicial: int, final: int, cantidad_piezas: int = 1):
        datos = self.obtener_guias_desde_excel(inicial, final)
        self.elegir_recolectar()

        textbox = self.page.locator("//*[@id='lblGuide']")
        for valor in datos:
            textbox.fill(f"{str(valor)}-{cantidad_piezas}")
            with self.page.expect_response("**/Home.aspx/FDApi", timeout=60000):
                textbox.press("Enter")
            textbox.wait_for(state="visible")
        self._take_screenshot("batch_collection_complete")

        # Pasos finales para completar la recolección
        self.page.get_by_role("button", name="Siguiente").click()
        
        # Llenar correo del cliente
        correo_textbox = self.page.get_by_role("textbox", name="Correo del cliente *")
        correo_textbox.wait_for(state="visible", timeout=30000)
        correo_textbox.fill("carlos.fernandez@forzalatam.com")
        
        # Dibujar en el canvas para activar botón Finalizar
        canvas = self.page.locator("canvas").first
        canvas.wait_for(state="visible")
        
        # Inyectar firma en el canvas para activar el botón Finalizar
        canvas = self.page.locator("canvas").first
        canvas.wait_for(state="visible")
        
        # Dibujar en el canvas con eventos reales (lo que funcionó visualmente)
        bbox = canvas.bounding_box()
        if bbox:
            self.page.mouse.move(bbox['x'] + 50, bbox['y'] + 100)
            self.page.mouse.down()
            for i in range(10):
                self.page.mouse.move(bbox['x'] + 50 + i * 15, bbox['y'] + 100 + (i % 3) * 10)
            self.page.mouse.up()
            self.page.wait_for_timeout(500)
        
        # Estrategia: Forzar el botón y ejecutar su acción directamente
        self.page.evaluate("""
            () => {
                const btn = document.querySelector('#btnFinish');
                if (!btn) return;
                
                // Habilitar el botón visualmente
                btn.disabled = false;
                btn.removeAttribute('disabled');
                btn.classList.remove('disabled');
                
                // Intentar obtener la función que ejecuta el botón
                // 1. Revisar si tiene un onclick directo
                if (btn.onclick) {
                    console.log('Ejecutando onclick directo');
                    btn.onclick();
                    return;
                }
            
                // 3. Si no encuentra nada, disparar un click event estándar
                console.log('Disparando evento click estándar');
                btn.dispatchEvent(new MouseEvent('click', {
                    view: window,
                    bubbles: true,
                    cancelable: true
                }));
            }
        """)
        
        # Esperar el mensaje de éxito
        self.page.get_by_role("alert").wait_for(state="visible", timeout=30000)
        expect(self.page.get_by_role("alert")).to_contain_text("Servicio completado exitosamente.")
        self._take_screenshot("recoleccion_finalizada")

    def elegir_recolectar(self):
        self.page.locator("//button[@type='button' and @data-type='Pickup' and contains(text(), 'Recolectar')]").first.click()
        self.page.locator("//button[@type='button' and contains(text(), 'Recolectar')]").click()

    @allure.step("Abrir recolección pendiente en posición {posicion}")
    def abrir_recoleccion_por_posicion(self, posicion: int):
        if posicion < 1:
            raise ValueError("La posición de la recolección debe ser mayor o igual a 1.")

        botones_recolectar = self.page.locator("//button[contains(normalize-space(.), 'Recolectar')]")
        botones_recolectar.nth(posicion - 1).click(timeout=30000)
        self.page.wait_for_load_state("networkidle")
        self._take_screenshot(f"pickup_position_{posicion}")

    @allure.step("Reportar visita fallida")
    def reportar_visita_fallida(self, ruta_imagen: str):
        self.page.get_by_role("button", name=re.compile("Visita fallida", re.I)).click(timeout=30000)
        self.page.wait_for_load_state("networkidle")

        self.seleccionar_primer_motivo_visita_fallida()
        self.cargar_imagen_visita_fallida(ruta_imagen)

        expect(self.page.get_by_text(re.compile("Ubicación detectada exitosamente", re.I))).to_be_visible(timeout=30000)
        reportar = self.page.locator("//button[contains(normalize-space(.), 'Reportar')]").last
        reportar.scroll_into_view_if_needed(timeout=10000)
        expect(reportar).to_be_enabled(timeout=120000)
        self.page.wait_for_function(
            """
            () => {
                const buttons = [...document.querySelectorAll('button')];
                const reportar = buttons.reverse().find((button) => button.textContent.includes('Reportar'));
                if (!reportar) return false;
                const style = window.getComputedStyle(reportar);
                return !reportar.disabled && style.visibility !== 'hidden' && style.display !== 'none';
            }
            """,
            timeout=120000,
        )
        reportar.click(timeout=120000)
        self._take_screenshot("failed_pickup_reported")

    def seleccionar_primer_motivo_visita_fallida(self):
        selected = self.page.evaluate(
            """
            () => {
                const select = document.querySelector('select');
                if (!select) return false;
                if (select.options.length > 1) {
                    select.selectedIndex = 1;
                }
                select.dispatchEvent(new Event('input', { bubbles: true }));
                select.dispatchEvent(new Event('change', { bubbles: true }));
                return Boolean(select.value);
            }
            """
        )
        if selected:
            return

        motivo_default = self.page.get_by_text("Información incompleta", exact=True)
        if motivo_default.count() > 0:
            return

        comboboxes = self.page.get_by_role("combobox")
        if comboboxes.count() > 0:
            try:
                comboboxes.first.select_option(index=1)
                return
            except Exception:
                pass

        selectores = self.page.locator("select")
        if selectores.count() > 0:
            opciones = selectores.first.locator("option")
            if opciones.count() > 1:
                selectores.first.select_option(index=1)
            else:
                selectores.first.select_option(index=0)
            return

        try:
            self.page.get_by_role("combobox").first.click(timeout=10000)
            self.page.get_by_role("option").nth(1).click(timeout=10000)
        except Exception:
            return

    def cargar_imagen_visita_fallida(self, ruta_imagen: str):
        ruta = Path(ruta_imagen)
        if not ruta.is_absolute():
            ruta = Path(__file__).resolve().parent.parent / ruta
        if not ruta.exists():
            raise FileNotFoundError(f"No existe la imagen para visita fallida: {ruta}")

        file_input = self.page.locator("input[type='file']").first
        file_input.set_input_files(str(ruta))
        self.page.wait_for_timeout(1000)

    @allure.step("Validar mensaje de visita fallida: '{mensaje}'")
    def validar_mensaje_visita_fallida(self, mensaje: str):
        expect(self.page.get_by_text(mensaje)).to_be_visible(timeout=120000)
        self._take_screenshot("failed_pickup_success_message")

    def obtener_guias_desde_excel(self, inicial: int, final: int) -> list:
        datos = self.leer_columna_excel(self.ruta_excel, self.hoja, self.columna)
        if inicial < 1:
            raise ValueError("El rango inicial debe ser mayor o igual a 1.")
        if final < inicial:
            raise ValueError("El rango final debe ser mayor o igual al rango inicial.")

        limite = min(len(datos), final)
        return datos[inicial - 1:limite]

    def leer_columna_excel(self, nombre_archivo: str, hoja: str, columna_letra: str) -> list:
        ruta_absoluta = os.path.abspath(nombre_archivo)
        if not os.path.exists(ruta_absoluta):
            raise FileNotFoundError(f"El archivo no existe: {ruta_absoluta}")

        try:
            wb = openpyxl.load_workbook(ruta_absoluta, data_only=True)
            if hoja not in wb.sheetnames:
                raise ValueError(f"La hoja '{hoja}' no existe.")
            
            ws = wb[hoja]
            col_idx = column_index_from_string(columna_letra)
            valores = []

            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                if row[0] is not None:
                    valores.append(row[0])
            return valores
        except Exception as e:
            print(f"Error leyendo el archivo Excel: {e}")
            return []

    # ==============================================================================
    # INTEGRACIÓN DE BASE DE DATOS (SQL Server con pyodbc)
    # ==============================================================================
    
    @allure.step("Obtener token OTP desde la Base de Datos")
    def obtener_token_db(self, telefono: str, entorno: str = None) -> str:
        """
        Conecta a SQL Server dinámicamente según el entorno, busca el token más reciente
        para un teléfono y lo devuelve.
        """
        entorno_normalizado = (entorno or "DEFAULT").upper()
        env_key = f"SQLSERVER_{entorno_normalizado}_CONNECTION_STRING"
        base_connection_string = os.getenv(env_key) or os.getenv("SQLSERVER_CONNECTION_STRING")

        if not base_connection_string:
            raise ValueError(
                f"No se encontró cadena de conexión. Configura {env_key} "
                "o SQLSERVER_CONNECTION_STRING en el archivo .env."
            )

        full_connection_string = base_connection_string.rstrip(";")
        if "Encrypt=" not in full_connection_string:
            full_connection_string += ";Encrypt=yes"
        if "TrustServerCertificate=" not in full_connection_string:
            full_connection_string += ";TrustServerCertificate=yes"
        full_connection_string += ";"

        try:
            # 4. Abrir la conexión
            conn = pyodbc.connect(full_connection_string)
            cursor = conn.cursor()

            # 5. La consulta SQL
            # IMPORTANTE: En pyodbc no se usan nombres como '@Telefono', se usan signos de interrogación '?'
            query = """
                SELECT TOP 1 srt.LoginToken
                FROM dbo.SenderReceiver sr
                INNER JOIN dbo.SenderReceiverLoginToken srt 
                    ON sr.ID = srt.SenderReceiverId
                WHERE sr.Phone = ?
                ORDER BY srt.DateCreated DESC
            """

            # 6. Equivalente a command.Parameters.Add() y ExecuteScalarAsync()
            # Pasamos 'telefono' como una tupla al método execute para prevenir SQL Injection
            cursor.execute(query, (telefono,))
            row = cursor.fetchone()

            # 7. Equivalente a result?.ToString()
            return str(row[0]) if row else ""

        except pyodbc.Error as e:
            print(f"Error de base de datos conectando a SQL Server: {e}")
            return ""
        except Exception as e:
            print(f"Error general en la obtención del token: {e}")
            return ""
        finally:
            # Buena práctica: Asegurarnos de cerrar la conexión en Python equivalente al 'await using'
            if 'conn' in locals():
                conn.close()

    # ==============================================================================
    # GENERADORES DE DATOS Y ARCHIVOS UTILS
    # ==============================================================================

    def guardar_guia_en_archivo(self, nombre_archivo: str, tipo_guia: str, guia: str):
        if not nombre_archivo.endswith(".txt"):
            nombre_archivo += ".txt"
            
        linea = f"Guia {tipo_guia} ----> {guia}\n"
        with open(nombre_archivo, "a", encoding="utf-8") as f:
            f.write(linea)

    def generar_nombre_qa(self) -> str:
        nombres = ["Carlos", "Ana", "Luis", "Sofía", "Pedro", "Laura", "Mario", "Paola", "Alfita", "Juancito", "Manuelito"]
        nombre_aleatorio = random.choice(nombres)
        numero = random.randint(100, 999)
        return f"QA {nombre_aleatorio}{numero}"

    def generar_telefono(self) -> str:
        primer_digito = str(random.randint(3, 7))
        resto = "".join(str(random.randint(0, 9)) for _ in range(7))
        return primer_digito + resto

    def generar_cuenta_aleatoria(self, longitud: int) -> str:
        return "".join(str(random.randint(0, 9)) for _ in range(longitud))

    # ==============================================================================
    # PAGO ZIGI (Paggo)
    # ==============================================================================

    @allure.step("Completar pago Zigi en link de pago")
    def completar_pago_zigi(self, link: str, nombre: str, apellido: str, email: str,
                             telefono: str, tarjeta: str, vencimiento: str, cvv: str,
                             dpi: str, departamento: str, direccion: str):
        self.page.goto(link, timeout=120000)
        self.page.wait_for_load_state("load")
        self.page.wait_for_timeout(1000)
        self._take_screenshot("zigi_pagina_pago_cargada")

        self.page.locator('input[name="nombre"]').wait_for(state="visible", timeout=60000)
        self._take_screenshot("zigi_formulario_visible")

        self.page.locator('input[name="nombre"]').click()
        self.page.locator('input[name="nombre"]').fill(nombre)
        self.page.locator('input[name="apellido"]').click()
        self.page.locator('input[name="apellido"]').fill(apellido)
        self.page.locator('input[name="email-field"]').click()
        self.page.locator('input[name="email-field"]').fill(email)
        self.page.locator('input[name="phone-field"]').click()
        self.page.locator('input[name="phone-field"]').fill(telefono)
        self._take_screenshot("zigi_datos_personales")
        self.page.wait_for_timeout(1000)

        self.page.locator('input[autocomplete="cc-number"]').click()
        self.page.locator('input[autocomplete="cc-number"]').fill(tarjeta)
        self.page.locator('input[autocomplete="cc-exp"]').click()
        self.page.locator('input[autocomplete="cc-exp"]').fill(vencimiento)
        self.page.locator('input[placeholder="123"]').click()
        self.page.locator('input[placeholder="123"]').fill(cvv)
        self._take_screenshot("zigi_datos_tarjeta")
        self.page.wait_for_timeout(1000)

        self.page.get_by_role("textbox", name="Ingrese su DPI o NIT").click()
        self.page.get_by_role("textbox", name="Ingrese su DPI o NIT").fill(dpi)
        self.page.get_by_role("button", name="Departamento*").first.click()
        self.page.get_by_role("option", name=departamento).click()
        self.page.wait_for_timeout(500)
        self.page.get_by_role("button", name=f"Departamento* {departamento}").nth(1).click()
        self.page.get_by_role("option", name=departamento).click()
        self.page.locator('input[name="direccionUno"]').click()
        self.page.locator('input[name="direccionUno"]').fill(direccion)
        self._take_screenshot("zigi_datos_facturacion")
        self.page.wait_for_timeout(1000)

        self.page.route("https://www.google.com/**", lambda route: route.abort())

        self.page.locator('[data-test-id="btn-loading-button"]').dblclick()

        self.page.get_by_text("Transaccion exitosa").wait_for(state="visible", timeout=60000)
        self._take_screenshot("zigi_transaccion_exitosa")

        self.page.get_by_text("Pago realizado exitosamente").wait_for(state="visible", timeout=10000)
        self._take_screenshot("zigi_comprobante_pago")

        print(f"  ✔ Pago Zigi completado — Transacción exitosa")
        allure.attach(
            f"Link: {link}\nNombre: {nombre} {apellido}\nResultado: Transaccion exitosa",
            name="Pago Zigi Completado",
            attachment_type=allure.attachment_type.TEXT,
        )

    # ==============================================================================
    # FLUJOS CORPORATIVOS Y EXEC
    # ==============================================================================

    @allure.step("Crear guías CORPORATIVO")
    def crear_guias_corp(self, datos):
        is_std = datos.tipo_guia == "Servicio Estándar"
        is_collet = str(datos.collet).lower() == "true"

        self.page.get_by_role("textbox", name="Selecciona un poblado").click()
        self.page.get_by_role("heading", name="Cahabon, Santa Maria Cahabon").click()
        self.page.get_by_role("button", name="calculator outline Calcular").click()

        if is_std:
            card = self.page.locator("div.card").filter(has_text="Servicio Estándar")
            card.get_by_role("button", name="Seleccionar").click()
        else:
            self.page.get_by_role("button", name="Seleccionar send").nth(1).click()

        self.page.get_by_role("radio").nth(1).click()
        self.page.get_by_role("textbox", name="Ingresar quién recibe").click()
        self.page.get_by_role("textbox", name="Ingresar quién recibe").fill("Juan")

        self.page.locator("input[name='ion-input-18']").click()
        self.page.locator("input[name='ion-input-18']").fill("Juan")
        self.page.locator("input[name='ion-input-19']").click()
        self.page.locator("input[name='ion-input-19']").fill("52452421")

        self.page.get_by_role("textbox", name="Ingresar dirección en destinatario").click()
        self.page.get_by_role("textbox", name="Ingresar dirección en destinatario").fill("Ciudad de Guatemala")
        self.page.get_by_role("button", name="Siguiente arrow forward").click()

        if datos.tipo_guia == "Servicio C.O.D.":
            self.page.get_by_role("textbox", name="Ingresar monto COD").click()
            self.page.get_by_role("textbox", name="Ingresar monto COD").fill("100")

        self.page.get_by_role("button", name="Siguiente arrow forward").click()

        if is_collet:
            self.page.locator("ion-col").filter(has_text=re.compile(r"^Crédito$")).locator("label").click()
            self.page.locator("ion-col").filter(has_text=re.compile(r"^Collect$")).locator("label").click()

        self.page.get_by_text("Arrow Back Regresar Mostrar").click()
        self.page.get_by_role("button", name="Mostrar Resumen arrow forward").click()
        self.page.wait_for_timeout(1000)

        heading_locator = self.page.locator("role=heading")
        matched_heading = heading_locator.filter(has_text="No. Guía Transporte FD").first

        if matched_heading.count() > 0:
            full_text = matched_heading.text_content()
            match = re.search(r"FD\d+", full_text)
            if match:
                guia = match.group(0)
                if is_std:
                    nombre_archivo = "guias_Corp_Collet_STD" if is_collet else "guias_Corp_Credito_STD"
                    tipo = "STD Collet" if is_collet else "STD Credito"
                else:
                    nombre_archivo = "guias_Corp_Collet_COD" if is_collet else "guias_Corp_Credito_COD"
                    tipo = "COD Collet" if is_collet else "COD Credito"
                print(f"Guía encontrada: {guia}")
                self.guardar_guia_en_archivo(nombre_archivo, tipo, guia)
                self._take_screenshot("corp_guide_created")
        else:
            print("No se encontró el heading con la guía.")

    @allure.step("Crear guías EXEC")
    def crear_guias_exec(self, datos):
        is_std = datos.tipo_guia == "Servicio Estándar"
        is_collet = str(datos.collet).lower() == "true"

        self.page.locator("ion-radio").filter(has_text="Clientes de Express Center").click()
        self.page.get_by_role("heading", name="Use el buscador para ver").click()
        self.page.wait_for_timeout(1000)
        self.page.get_by_role("textbox", name="Selecciona un poblado").click()
        self.page.get_by_role("heading", name="Cahabon, Santa Maria Cahabon").click()
        self.page.get_by_role("button", name="calculator outline Calcular").click()
        self.page.wait_for_timeout(1000)

        if is_std:
            card = self.page.locator("div.card").filter(has_text="Servicio Estándar")
            card.get_by_role("button", name="Seleccionar").click()
        else:
            card = self.page.locator("div.card").filter(has_text="Servicio C.O.D.")
            card.get_by_role("button", name="Seleccionar").click()

        nombre_envia = self.generar_nombre_qa()
        telefono_envia = self.generar_telefono()

        self.page.get_by_placeholder("Ingresar nombre de contacto").first.click()
        self.page.get_by_placeholder("Ingresar nombre de contacto").first.fill(nombre_envia)
        self.page.get_by_placeholder("Ingresar teléfono").first.click()
        self.page.get_by_placeholder("Ingresar teléfono").first.fill(telefono_envia)
        self.page.get_by_role("textbox", name="Ingresar correo electrónico").first.click()
        self.page.get_by_role("textbox", name="Ingresar correo electrónico").first.fill("carlos.fernandez@forzalatam.com")

        nombre_recibe = self.generar_nombre_qa()
        self.page.get_by_placeholder("Ingresar nombre de contacto").nth(1).click()
        self.page.get_by_placeholder("Ingresar nombre de contacto").nth(1).fill(nombre_recibe)

        telefono_recibe = self.generar_telefono()
        self.page.get_by_placeholder("Ingresar teléfono").nth(1).click()
        self.page.get_by_placeholder("Ingresar teléfono").nth(1).fill(telefono_recibe)

        self.page.get_by_role("textbox", name="Ingresar dirección en destinatario").click()
        self.page.get_by_role("textbox", name="Ingresar dirección en destinatario").fill("Ciudad de Guatemala")

        self.page.get_by_role("radio").nth(1).click()
        self.page.get_by_role("button", name="Siguiente arrow forward").click()

        if datos.tipo_guia == "Servicio C.O.D.":
            self.page.get_by_role("textbox", name="Ingresar monto COD").click()
            self.page.get_by_role("textbox", name="Ingresar monto COD").fill("100")
            self.page.get_by_role("textbox", name="Seleccione Banco").click()
            self.page.get_by_role("heading", name="BANRURAL - BANCO DE").click()
            self.page.get_by_role("textbox", name="Seleccione Tipo de Cuenta").click()
            self.page.get_by_role("heading", name="Ahorro").click()
            self.page.get_by_role("textbox", name="Ingresar número de cuenta").click()
            cuenta = self.generar_cuenta_aleatoria(15)
            self.page.get_by_role("textbox", name="Ingresar número de cuenta").fill(cuenta)
            self.page.get_by_role("textbox", name="Ingresar nombre de la cuenta").click()
            self.page.get_by_role("textbox", name="Ingresar nombre de la cuenta").fill("QA")
            self.page.get_by_role("textbox", name="Ingresar Documento de").click()
            self.page.get_by_role("textbox", name="Ingresar Documento de").fill("6352452")

        self.page.get_by_role("button", name="Siguiente arrow forward").click()

        if is_collet:
            self.page.locator("ion-col").filter(has_text=re.compile(r"^Collect$")).locator("label").click()

        self.page.get_by_text("Arrow Back Regresar Mostrar").click()
        self.page.get_by_role("button", name="Mostrar Resumen arrow forward").click()
        self.page.wait_for_timeout(1000)

        heading_locator = self.page.locator("role=heading")
        matched_heading = heading_locator.filter(has_text="No. Guía Transporte FD").first

        if matched_heading.count() > 0:
            full_text = matched_heading.text_content()
            match = re.search(r"FD\d+", full_text)
            if match:
                guia = match.group(0)
                if is_std:
                    nombre_archivo = "guias_EXEC_Collet_STD" if is_collet else "guias_Corp_Credito_STD"
                    tipo = "STD Collet" if is_collet else "STD Credito"
                else:
                    nombre_archivo = "guias_EXEC_Collet_COD" if is_collet else "guias_Corp_Credito_COD"
                    tipo = "COD Collet" if is_collet else "COD Credito"
                print(f"Guía encontrada: {guia}")
                self.guardar_guia_en_archivo(nombre_archivo, tipo, guia)
                self._take_screenshot("exec_guide_created")
        else:
            print("No se encontró el heading con la guía.")
